import os
import re
import csv
import logging
from pathlib import Path
from typing import List, Set, Tuple, Optional
import openpyxl
import pandas as pd
from pypdf import PdfReader, PdfWriter

import configuracion as config
from modules.dominio.modelos import RegistroDocumento, MetadatosArchivo
from modules.aplicacion.puertos import RepositorioExcel, ServicioPdf, ServicioAlmacenamiento

logger = logging.getLogger(__name__)

# --- Helper functions originally from folder_builder ---
_CARACTERES_INVALIDOS_RE = re.compile(r'[\\/:*?"<>|]')
_TABLA_ROMANA = [
    ('M', 1000), ('CM', 900), ('D', 500), ('CD', 400),
    ('C', 100),  ('XC', 90),  ('L', 50),  ('XL', 40),
    ('X', 10),   ('IX', 9),   ('V', 5),   ('IV', 4),  ('I', 1),
]

def _romano_a_arabigo(s: str) -> str:
    texto = s.strip().upper()
    if not texto:
        return s
    resultado = 0
    i = 0
    for numeral, valor in _TABLA_ROMANA:
        while texto[i:i + len(numeral)] == numeral:
            resultado += valor
            i += len(numeral)
    if i == len(texto) and resultado > 0:
        return str(resultado)
    return s

_NOMBRES_MESES = {
    1:  "1. ENERO",
    2:  "2. FEBRERO",
    3:  "3. MARZO",
    4:  "4. ABRIL",
    5:  "5. MAYO",
    6:  "6. JUNIO",
    7:  "7. JULIO",
    8:  "8. AGOSTO",
    9:  "9. SEPTIEMBRE",
    10: "10. OCTUBRE",
    11: "11. NOVIEMBRE",
    12: "12. DICIEMBRE",
}

def _sanitizar(nombre: str, por_defecto: str = "Sin_Nombre") -> str:
    if not nombre or not str(nombre).strip():
        return por_defecto
    limpio = _CARACTERES_INVALIDOS_RE.sub("_", str(nombre).strip())
    limpio = limpio.rstrip(". ")
    return limpio or por_defecto

def _primer_nombre_completo(original: str) -> str:
    if not original or not str(original).strip():
        return ""
    partes = str(original).split(",", 1)
    return partes[0].strip()

def _parsear_fecha(fecha_str: str) -> tuple[Optional[int], Optional[int]]:
    try:
        partes = str(fecha_str).strip().split("/")
        if len(partes) == 3:
            dia, mes, anio = int(partes[0]), int(partes[1]), int(partes[2])
            return anio, mes
    except (ValueError, IndexError):
        pass
    return None, None


class RepositorioExcelPandas(RepositorioExcel):
    def cargar_metadatos(self, ruta_excel: str) -> MetadatosArchivo:
        ruta = Path(ruta_excel)
        if not ruta.exists():
            raise FileNotFoundError(f"Archivo Excel no encontrado: {ruta}")

        libro = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
        hoja = libro.active

        siglo_orig = str(hoja.cell(config.META_ROW_SIGLO,  1).value or "").strip()
        acervo_orig = str(hoja.cell(config.META_ROW_ACERVO, 1).value or "").strip()
        libro.close()

        # Extraer siglo: «Seccion: XVI» → 'XVI'
        m_siglo = re.search(r':\s*(.+)$', siglo_orig)
        siglo = m_siglo.group(1).strip() if m_siglo else siglo_orig or "Sin_Siglo"

        # Extraer acervo: «Código del fondo: N7» → '7'
        m_num = re.search(r'N(\d+)', acervo_orig, re.IGNORECASE)
        if m_num:
            num_acervo = m_num.group(1)
        else:
            m_colon = re.search(r':\s*(.+)$', acervo_orig)
            num_acervo = m_colon.group(1).strip() if m_colon else acervo_orig or "Sin_Num"

        logger.info(f"Metadatos Excel cargados -> Siglo: {siglo} | Acervo: {num_acervo}")
        return MetadatosArchivo(
            siglo=siglo,
            num_acervo=num_acervo,
            siglo_original=siglo_orig,
            acervo_original=acervo_orig
        )

    def cargar_registros(self, ruta_excel: str, filas_omitidas: int) -> List[RegistroDocumento]:
        ruta = Path(ruta_excel)
        if not ruta.exists():
            raise FileNotFoundError(f"Archivo Excel no encontrado: {ruta}")

        df = pd.read_excel(
            ruta,
            skiprows=filas_omitidas,
            header=0,
            engine="openpyxl",
            dtype=str,
        )

        df.columns = [str(c).strip() for c in df.columns]
        
        # Eliminar columnas sin nombre
        columnas_anonimas = [c for c in df.columns if str(c).startswith("Unnamed:")]
        if columnas_anonimas:
            df.drop(columns=columnas_anonimas, inplace=True)

        # Ajuste de DATA CRÓNICA
        mapeo_nombres = {}
        columnas_cronicas = [c for c in df.columns if "DATA CR" in c.upper()]
        if len(columnas_cronicas) >= 2:
            mapeo_nombres[columnas_cronicas[0]] = "FECHA INICIAL"
            mapeo_nombres[columnas_cronicas[1]] = "FECHA FINAL"
        elif len(columnas_cronicas) == 1:
            mapeo_nombres[columnas_cronicas[0]] = "FECHA INICIAL"
        df.rename(columns=mapeo_nombres, inplace=True)

        # Indexación real (1-based + filas_omitidas + offset de cabecera)
        df.index = df.index + filas_omitidas + 2
        df.dropna(how="all", inplace=True)
        df = df.fillna("")

        # Filtro de marcadores de sección
        primera_columna = df.columns[0]
        _patron_marcador_seccion = re.compile(r'^\s*(protocolo|registro)\b', re.IGNORECASE)

        def _es_fila_de_datos(val: str) -> bool:
            v = str(val).strip()
            if not v or _patron_marcador_seccion.match(v):
                return False
            return True

        df = df[df[primera_columna].apply(_es_fila_de_datos)]

        registros = []
        for idx, row in df.iterrows():
            registros.append(
                RegistroDocumento(
                    fila_excel=idx,
                    registro_id=str(row.get(config.COL_REGISTRO, "")).strip(),
                    escribano=str(row.get(config.COL_ESCRIBANO, "")).strip(),
                    protocolo=str(row.get(config.COL_PROTOCOLO, "")).strip(),
                    folios_origen=str(row.get(config.COL_FOLIOS, "")).strip(),
                    titulo=str(row.get(config.COL_TITULO_EST, "")).strip(),
                    fecha_inicio=str(row.get(config.COL_FECHA_INI, "")).strip(),
                    fecha_fin=str(row.get(config.COL_FECHA_FIN, "")).strip(),
                    interesado1=str(row.get(config.COL_INT1, "")).strip(),
                    interesado2=str(row.get(config.COL_INT2, "")).strip(),
                    lugar=str(row.get(config.COL_TOPICA, "")).strip(),
                    observaciones=str(row.get(config.COL_OBS, "")).strip() if config.COL_OBS in row else ""
                )
            )

        return registros


class ServicioPdfPyPdf(ServicioPdf):
    def __init__(self):
        self._lectores = {}

    def _obtener_lector(self, ruta_pdf: str) -> PdfReader:
        if ruta_pdf not in self._lectores:
            ruta = Path(ruta_pdf)
            if not ruta.exists():
                raise FileNotFoundError(f"PDF no encontrado: {ruta}")
            self._lectores[ruta_pdf] = PdfReader(str(ruta))
        return self._lectores[ruta_pdf]

    def obtener_cantidad_paginas(self, ruta_pdf: str) -> int:
        lector = self._obtener_lector(ruta_pdf)
        return len(lector.pages)

    def extraer_paginas(
        self,
        ruta_pdf: str,
        numeros_paginas: List[int],
        ruta_destino: str,
        simulacion: bool = False,
        paginas_ignoradas: Set[int] = None,
    ) -> bool:
        lector = self._obtener_lector(ruta_pdf)
        total_paginas = len(lector.pages)
        paginas_ignoradas = paginas_ignoradas or set()
        escritor = PdfWriter()
        ignoradas_omitidas = []

        for num_pag in numeros_paginas:
            if num_pag in paginas_ignoradas:
                ignoradas_omitidas.append(num_pag)
                continue

            idx = num_pag - 1
            if idx < 0 or idx >= total_paginas:
                logger.error(f"Página {num_pag} fuera de rango. Fallo en extracción.")
                return False
            escritor.add_page(lector.pages[idx])

        if ignoradas_omitidas:
            logger.info(f"Páginas ignoradas: {ignoradas_omitidas}")

        if len(escritor.pages) == 0:
            logger.warning(f"Todas las páginas de '{Path(ruta_destino).name}' fueron ignoradas. No se escribe archivo.")
            return True

        destino = Path(ruta_destino)
        if simulacion:
            logger.info(f"[SIMULACION] Se escribiría: {destino}")
            return True

        destino.parent.mkdir(parents=True, exist_ok=True)
        with open(destino, "wb") as f:
            escritor.write(f)

        logger.debug(f"PDF escrito: {destino}")
        return True


class ServicioAlmacenamientoWindows(ServicioAlmacenamiento):
    NOMBRES_COLUMNAS_PENDIENTES = [
        "N° DE REGISTRO", "ESCRIBANO/NOTARIO", "N° DE FOLIOS",
        "TITULO/ESCRITURA", "INTERESADO 1", "INTERESADO 2",
        "MOTIVO",
    ]

    def construir_ruta_destino(
        self,
        directorio_salida: str,
        metadatos: MetadatosArchivo,
        registro: RegistroDocumento,
        simulacion: bool = False,
    ) -> str:
        anio, mes = _parsear_fecha(registro.fecha_inicio)
        anio_str  = str(anio) if anio else "Sin_Año"
        mes_str   = _NOMBRES_MESES.get(mes, "Sin_Mes") if mes else "Sin_Mes"

        prot_str = _sanitizar(registro.protocolo, "Sin_Protocolo")
        reg_str  = _sanitizar(registro.registro_id, "Sin_Registro")

        int1_primer = _primer_nombre_completo(registro.interesado1)
        int2_primer = _primer_nombre_completo(registro.interesado2)

        siglo_disp = _romano_a_arabigo(_sanitizar(metadatos.siglo, "Sin_Siglo"))

        ruta_carpeta = (
            Path(directorio_salida)
            / f"ACERVO DOCUMENTAL NUMERO {_sanitizar(metadatos.num_acervo, 'Sin_Acervo')}"
            / f"SIGLO {siglo_disp}"
            / "FONDO DOCUMENTAL"
            / _sanitizar(registro.escribano, "Sin_Escribano")
            / anio_str
            / f"PROTOCOLO {prot_str}"
            / f"REGISTRO {reg_str}"
            / _sanitizar(registro.titulo, "Sin_Titulo")
            / mes_str
            / _sanitizar(int1_primer, "Sin_Interesado1")
        )

        base_nombre = _sanitizar(int2_primer, "Sin_Interesado2")
        ruta_destino = ruta_carpeta / f"{base_nombre}.pdf"

        # Resolver colisión
        if ruta_destino.exists() and not simulacion:
            contador = 2
            while True:
                candidato = ruta_carpeta / f"{base_nombre}_{contador}.pdf"
                if not candidato.exists():
                    ruta_destino = candidato
                    break
                contador += 1

        if not simulacion:
            ruta_carpeta.mkdir(parents=True, exist_ok=True)

        return str(ruta_destino)

    def registrar_pendiente_csv(
        self,
        ruta_csv: str,
        registro: RegistroDocumento,
        motivo: str,
        simulacion: bool = False,
    ) -> None:
        if simulacion:
            logger.info(f"[SIMULACION] Pendiente: Fila {registro.fila_excel} - {motivo}")
            return

        destino = Path(ruta_csv)
        destino.parent.mkdir(parents=True, exist_ok=True)
        escribir_cabecera = destino.stat().st_size == 0 if destino.exists() else True

        with open(destino, "a", newline="", encoding="utf-8-sig") as f:
            escritor = csv.DictWriter(f, fieldnames=self.NOMBRES_COLUMNAS_PENDIENTES, extrasaction="ignore")
            if escribir_cabecera:
                escritor.writeheader()
            
            dicc_fila = registro.a_diccionario()
            dicc_fila["MOTIVO"] = motivo
            escritor.writerow(dicc_fila)
